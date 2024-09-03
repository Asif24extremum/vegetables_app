import os
import time
import pandas as pd
import streamlit as st
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
    ElementNotInteractableException,
    ElementClickInterceptedException
)
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Define the folder where all data will be saved
output_folder = 'scraped_data'
os.makedirs(output_folder, exist_ok=True)

# Initialize session state variables
if 'stop_scraping' not in st.session_state:
    st.session_state.stop_scraping = False

if 'download_files' not in st.session_state:
    st.session_state.download_files = {}


def clear_previous_data():
    """Clear the previous scraped data and session state."""
    if os.path.exists(output_folder):
        for file in os.listdir(output_folder):
            file_path = os.path.join(output_folder, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
    st.session_state.download_files.clear()


def append_to_excel(df, file_path):
    """Append DataFrame to an Excel file using openpyxl to ensure proper alignment."""
    if not os.path.exists(file_path):
        df.to_excel(file_path, index=False)
    else:
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        for sheetname in writer.sheets:
            df.to_excel(writer, sheet_name=sheetname, index=False, header=False,
                        startrow=writer.sheets[sheetname].max_row)
        writer.save()


def scrape_agmarknet(driver, search_terms):
    if st.session_state.stop_scraping:
        return pd.DataFrame()

    url = 'https://agmarknet.gov.in'
    driver.get(url)
    time.sleep(5)

    try:
        element_present = EC.element_to_be_clickable(
            (By.XPATH, "//td[text()='Vegetables']/preceding-sibling::td/input[@type='image']"))
        WebDriverWait(driver, 10).until(element_present)
    except TimeoutException:
        print("Timed out waiting for page to load 'Vegetables' section.")
        return pd.DataFrame()

    vegetables_section = driver.find_element(By.XPATH,
                                             "//td[text()='Vegetables']/preceding-sibling::td/input[@type='image']")
    vegetables_section.click()
    time.sleep(2)

    data = []
    seen_items = set()

    def get_vegetable_items():
        return driver.find_elements(By.XPATH, "//table[@title='Vegetables']//tr[td/input[@type='image']]")

    def click_and_collect_details(index, retry_count=3):
        if st.session_state.stop_scraping:
            return

        for attempt in range(retry_count):
            try:
                if st.session_state.stop_scraping:
                    return

                vegetable_items = get_vegetable_items()
                item = vegetable_items[index]
                veg_name = item.find_elements(By.TAG_NAME, "td")[1].text

                if veg_name in seen_items:
                    return

                plus_button = item.find_element(By.XPATH, "./td[1]/input[@type='image']")
                plus_button.click()

                time.sleep(2)

                expanded_details_table = driver.find_element(By.XPATH,
                                                             f"//tr[td[text()='{veg_name}']]/following-sibling::tr[1]//table")
                expanded_details = expanded_details_table.find_elements(By.TAG_NAME, "td")
                Search = 'N/A'
                if expanded_details:
                    for i in range(0, len(expanded_details), 4):
                        variety = expanded_details[i].text
                        max_price = expanded_details[i + 1].text
                        min_price = expanded_details[i + 2].text
                        modal_price = expanded_details[i + 3].text
                        data.append([Search, veg_name, variety, max_price, min_price, modal_price])
                    seen_items.add(veg_name)
                    print(f"Collected data for {veg_name}")
                break

            except Exception as e:
                print(f"Error processing {veg_name}: {e}")

    vegetable_items = get_vegetable_items()
    for index in range(len(vegetable_items)):
        if st.session_state.stop_scraping:
            break
        click_and_collect_details(index)

    df = pd.DataFrame(data, columns=['Search Term', 'Agmarknet_Commodity', 'Agmarknet_Variety', 'Agmarknet_MAX',
                                     'Agmarknet_MIN', 'Agmarknet_Modal'])
    df['Source'] = 'Agmarknet'
    file_path = os.path.join(output_folder, 'agmarknet_vegetable_prices.xlsx')
    append_to_excel(df, file_path)
    return df, file_path


def scrape_bigbasket(driver, search_terms):
    if st.session_state.stop_scraping:
        return pd.DataFrame()

    url = 'https://www.bigbasket.com/'
    driver.get(url)
    time.sleep(5)

    data = []

    def save_page_source(term):
        with open(f"error_page_{term}.html", "w", encoding="utf-8") as file:
            file.write(driver.page_source)

    def get_dropdown_prices():
        if st.session_state.stop_scraping:
            return 'N/A'

        try:
            wait = WebDriverWait(driver, 10)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'ul[role="listbox"]')))
            time.sleep(2)
            dropdown_elements = driver.find_elements(By.CSS_SELECTOR,
                                                     'ul[role="listbox"] li div.PackChanger___StyledDiv-sc-newjpv-4')
            dropdown_prices = []
            for elem in dropdown_elements:
                if st.session_state.stop_scraping:
                    return 'N/A'
                size_info = elem.find_element(By.CSS_SELECTOR, 'div.w-3\\/4').text.strip()
                price_info = elem.find_element(By.CSS_SELECTOR,
                                               'span.PackChanger___StyledLabel4-sc-newjpv-6').text.strip()
                dropdown_prices.append(f"{size_info}: {price_info}")
            return ', '.join(dropdown_prices) if dropdown_prices else 'N/A'
        except Exception as e:
            print(f"Failed to get dropdown prices: {e}")
            return 'N/A'

    for term in search_terms:
        if st.session_state.stop_scraping:
            break

        try:
            print(f"Searching for term: {term}")
            wait = WebDriverWait(driver, 20)
            search_bar = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder="Search for Products..."]')))
            search_bar.clear()
            search_bar.send_keys(term)
            search_bar.send_keys(Keys.RETURN)

            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.SKUDeck___StyledDiv-sc-1e5d9gk-0')))
            time.sleep(5)

            product_cards = driver.find_elements(By.CSS_SELECTOR, 'div.SKUDeck___StyledDiv-sc-1e5d9gk-0')
            print(f"Found {len(product_cards)} product cards for term: {term}")

            for card in product_cards[:4]:
                if st.session_state.stop_scraping:
                    break

                try:
                    brand_element = card.find_element(By.CSS_SELECTOR, 'span.BrandName___StyledLabel2-sc-hssfrl-1')
                    product_element = card.find_element(By.CSS_SELECTOR, 'h3.block')
                    title = f"{brand_element.text} {product_element.text}"

                    price_element = card.find_element(By.CSS_SELECTOR, 'span.Pricing___StyledLabel-sc-pldi2d-1')
                    price = price_element.text

                    original_price_element = card.find_element(By.CSS_SELECTOR,
                                                               'span.Pricing___StyledLabel2-sc-pldi2d-2')
                    original_price = original_price_element.text if original_price_element else 'N/A'

                    discount_element = card.find_element(By.CSS_SELECTOR, 'span.Tags___StyledLabel2-sc-aeruf4-1')
                    discount = discount_element.text if discount_element else 'N/A'

                    pack_sizes = card.find_elements(By.CSS_SELECTOR, 'span.PackChanger___StyledLabel-sc-newjpv-1')
                    if pack_sizes:
                        for size in pack_sizes:
                            if st.session_state.stop_scraping:
                                break
                            size_text = size.text
                            actions = webdriver.ActionChains(driver)
                            actions.move_to_element(size).click().perform()
                            dropdown_prices = get_dropdown_prices()
                            data.append({
                                'Search Term': term,
                                'BigBasket_Title': title,
                                'BigBasket_Price': price,
                                'BigBasket_Original_Price': original_price,
                                'BigBasket_Discount': discount,
                                'BigBasket_Pack_Size': size_text,
                                'BigBasket_Dropdown_Prices': dropdown_prices
                            })
                            print(f"Appended data for product: {title} with size {size_text}")
                    else:
                        data.append({
                            'Search Term': term,
                            'BigBasket_Title': title,
                            'BigBasket_Price': price,
                            'BigBasket_Original_Price': original_price,
                            'BigBasket_Discount': discount,
                            'BigBasket_Pack_Size': 'N/A',
                            'BigBasket_Dropdown_Prices': 'N/A'
                        })
                        print(f"Appended data for product: {title} with no dropdown")

                except Exception as e:
                    print(f"Error processing a product card: {e}")
                    continue

        except Exception as e:
            print(f"Failed to search or extract data for term '{term}': {e}")
            save_page_source(term)

    df = pd.DataFrame(data)
    df['Source'] = 'BigBasket'
    file_path = os.path.join(output_folder, 'bigbasket_Products_price.xlsx')
    append_to_excel(df, file_path)
    return df, file_path


def scrape_dmart(driver, search_terms):
    if st.session_state.stop_scraping:
        return pd.DataFrame()

    url = "https://www.dmart.in"
    driver.get(url)
    time.sleep(5)

    try:
        pincode_popup = driver.find_element(By.CLASS_NAME, "pincode-widget_pincode-header__bR5DG")
        pincode_input = pincode_popup.find_element(By.ID, "pincodeInput")
        pincode_input.send_keys("122001, Gurgaon")
        time.sleep(2)

        first_result = driver.find_element(By.CSS_SELECTOR,
                                           "ul.pincode-widget_pincode-list___pWVx li.pincode-widget_pincode-item__qsZwZ button")
        first_result.click()

        time.sleep(5)
        confirm_button = driver.find_element(By.XPATH, "//button[text()='CONFIRM LOCATION']")
        confirm_button.click()
        time.sleep(5)

        all_data = []

        for term in search_terms:
            if st.session_state.stop_scraping:
                break

            attempts = 0
            max_attempts = 3
            success = False

            while attempts < max_attempts and not success:
                if st.session_state.stop_scraping:
                    break

                try:
                    search_input = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "scrInput"))
                    )
                    search_input.clear()
                    search_input.send_keys(term)

                    search_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.search_searchButton__J9wVN"))
                    )
                    search_button.click()

                    time.sleep(5)

                    product_card_html = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.vertical-card_card-vertical__Q8seS"))
                    ).get_attribute('outerHTML')

                    soup = BeautifulSoup(product_card_html, 'html.parser')

                    title_elem = soup.find('div', class_='vertical-card_title__pMGg9')
                    mrp_elem = soup.find('span', style="text-decoration: line-through;")
                    dmart_price_elem = soup.find_all('span', class_='vertical-card_amount__80Zwk')
                    offer_elem = soup.find('div', class_='vertical-card_section-right__4rjsN')

                    title = title_elem.text.strip() if title_elem else 'N/A'
                    mrp = mrp_elem.text.strip() if mrp_elem else 'N/A'
                    dmart_price = dmart_price_elem[1].text.strip() if len(dmart_price_elem) > 1 else 'N/A'
                    offer = offer_elem.text.strip() if offer_elem else 'N/A'

                    dropdown_data = []
                    dropdown = soup.find('div', class_='MuiFormControl-root')
                    if dropdown:
                        try:
                            dropdown_element = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.ID, "demo-customized-select"))
                            )
                            dropdown_element.click()
                            time.sleep(2)

                            dropdown_options = WebDriverWait(driver, 10).until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "ul.MuiMenu-list li"))
                            )
                            for option in dropdown_options:
                                if st.session_state.stop_scraping:
                                    break
                                weight_elem = option.find_element(By.CSS_SELECTOR, "span[style='padding-left: 0px;']")
                                price_elem = option.find_element(By.CSS_SELECTOR,
                                                                 "span.bootstrap-select_infoTxt-value__kT4zZ")
                                weight = weight_elem.text.strip() if weight_elem else 'N/A'
                                price = price_elem.text.strip() if price_elem else 'N/A'
                                dropdown_data.append(f'{weight}: {price}')

                            driver.find_element(By.CSS_SELECTOR, "body").click()
                            time.sleep(1)
                        except NoSuchElementException:
                            print(f"No dropdown options found for {title}.")
                        except Exception as e:
                            print(f"An error occurred while handling the dropdown for {title}: {e}")

                    all_data.append({
                        'Search Term': term,
                        'DMart_Title': title,
                        'DMart_MRP': mrp,
                        'DMart_Price': dmart_price,
                        'DMart_Offer': offer,
                        'DMart_Dropdown_Options': ', '.join(dropdown_data)
                    })

                    success = True

                except Exception as e:
                    print(f"An error occurred while searching for '{term}': {e}")
                    attempts += 1
                    time.sleep(3)

        df = pd.DataFrame(all_data)
        df['Source'] = 'DMart'
        file_path = os.path.join(output_folder, 'dmart_product_data.xlsx')
        append_to_excel(df, file_path)
        return df, file_path

    except Exception as e:
        print(f"An error occurred: {e}")
        return pd.DataFrame(), None


def scrape_hyperpure(driver, search_terms):
    if st.session_state.stop_scraping:
        return pd.DataFrame()

    url = "https://www.hyperpure.com/in/fruits-vegetables?&type=CATALOG&cheapestProduct=0&discountedProduct=0&entity_id=&entity_type=&parent_reference_id=96887735-46cc-4fdb-8d19-65387afdc926-1721711561231890664&parent_reference_type=&search_source=&source_page=&sub_reference_id=&sub_reference_type="
    driver.get(url)
    time.sleep(5)

    all_data = []

    def scrape_data(search_term):
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        products = soup.find_all('div', class_='CatalogCard_catalogCard__mGd27')
        data = []
        for product in products:
            if st.session_state.stop_scraping:
                break
            try:
                product_title = product.find('div',
                                             class_='my-2 word-break text-align-left w-600 fs-16 CatalogCard_truncate__dW5IB').text.strip()
                price = product.find('span', class_='w-800 text-gray-900 CatalogCard_price__Pf25D').text.strip()
                category = product_title.split(",")[0]

                supersaver_info = product.find('div', class_='CatalogCard_offerTag__7QmgG')
                if supersaver_info:
                    supersaver_info = ' | '.join(
                        [offer.text.strip() for offer in product.find_all('div', class_='CatalogCard_offerV2__V6o1z')])
                else:
                    supersaver_info = "N/A"

                data.append({
                    'Search Term': search_term,
                    'Hyperpure_Product_Title': product_title,
                    'Hyperpure_Price': price,
                    'Hyperpure_Category': category,
                    'Hyperpure_SUPERSAVER_Information': supersaver_info
                })
            except AttributeError:
                continue

        return data

    for term in search_terms:
        if st.session_state.stop_scraping:
            break
        try:
            print(f"Searching for {term}...")
            search_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input.SearchInput_searchInput__8P47H'))
            )
            search_input.clear()
            search_input.send_keys(term)

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, '#react-autowhatever-1 .SearchInput_suggestionsList__dx_Xc'))
            )
            first_suggestion = driver.find_element(By.CSS_SELECTOR, '#react-autowhatever-1--item-0')
            first_suggestion.click()

            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'CatalogCard_catalogCard__mGd27'))
            )

            all_data.extend(scrape_data(term))

        except TimeoutException:
            print(f"No results found for {term} or the page took too long to load.")
        except NoSuchElementException:
            print(f"No search suggestions found for {term}.")

    df = pd.DataFrame(all_data)
    df['Source'] = 'Hyperpure'
    file_path = os.path.join(output_folder, 'hyperpure_product_data.xlsx')
    append_to_excel(df, file_path)
    return df, file_path


def scrape_jiomart(driver, search_terms):
    if st.session_state.stop_scraping:
        return pd.DataFrame()

    url = 'https://www.jiomart.com/'
    driver.get(url)
    time.sleep(5)

    try:
        location_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'btn_pin_code_delivery'))
        )
        location_button.click()

        enter_pincode_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'btn_enter_pincode'))
        )
        enter_pincode_button.click()

        pin_code_input = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, 'rel_pincode'))
        )
        pin_code_input.clear()
        pin_code_input.send_keys('122001')

        apply_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'btn_pincode_submit'))
        )
        apply_button.click()

        time.sleep(5)
        delivery_location = driver.find_element(By.ID, 'delivery_city_pincode').text
        if '122001' in delivery_location:
            print("Location set successfully!")
        else:
            print("Failed to set the location.")

        all_results_df = pd.DataFrame(columns=[
            'Search Term',
            'JioMart_Title',
            'JioMart_Offer',
            'JioMart_Price',
            'JioMart_Real_Price'
        ])

        for term in search_terms:
            if st.session_state.stop_scraping:
                break

            print(f"Searching for '{term}'...")
            driver.get(url)
            time.sleep(5)

            search_input = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.ID, 'autocomplete-0-input'))
            )
            search_input.clear()
            search_input.send_keys(term)
            search_input.send_keys(Keys.RETURN)

            time.sleep(10)

            try:
                first_product_card = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, '.plp-card-wrapper'))
                )
                product_card_html = first_product_card.get_attribute('outerHTML')

                soup = BeautifulSoup(product_card_html, 'html.parser')

                title = soup.find('div', class_='plp-card-details-name').text.strip()
                offer = soup.find('div', class_='plp-card-details-discount')
                offer_text = offer.text.strip() if offer else 'No offer'

                price = soup.find('span', class_='jm-heading-xxs').text.strip()
                real_price = soup.find('span', class_='jm-body-xxs').text.strip()

                print(f"Title: {title}")
                print(f"Offer: {offer_text}")
                print(f"Price: {price}")
                print(f"Real Price: {real_price}")

                data = {
                    'Search Term': [term],
                    'JioMart_Title': [title],
                    'JioMart_Offer': [offer_text],
                    'JioMart_Price': [price],
                    'JioMart_Real_Price': [real_price]
                }
                df = pd.DataFrame(data)

                all_results_df = pd.concat([all_results_df, df], ignore_index=True)

            except Exception as e:
                print(f"Failed to extract data for '{term}': {e}")

        all_results_df['Source'] = 'JioMart'  # Add the source column

        excel_file = os.path.join(output_folder, 'jiomart_product_data.xlsx')
        append_to_excel(all_results_df, excel_file)
        return all_results_df, excel_file
    except Exception as e:
        print("Exception in jiomart DATA", e)
        return pd.DataFrame(), None


def main(selected_websites, search_terms):
    # Initialize stop_scraping flag
    st.session_state.stop_scraping = False

    # Clear previous data
    clear_previous_data()

    # Initialize WebDriver
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1200')

    # Use the exact path provided in the error log
    service = Service('/home/appuser/.wdm/drivers/chromedriver/linux64/114.0.5735.90/chromedriver')

    # Initialize WebDriver with the specified service and options
    driver = webdriver.Chrome(service=service, options=options)

    # Use ChromeDriverManager to download and use the correct version of ChromeDriver
    # driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
    
    try:
        # Dictionary to store data from selected websites
        all_data = {}

        # Define the columns for the master DataFrame
        columns = ['Search Term', 'JioMart_Title', 'JioMart_Offer', 'JioMart_Price', 'JioMart_Real_Price', 'Source',
                   'DMart_Title', 'DMart_MRP', 'DMart_Price', 'DMart_Offer', 'DMart_Dropdown_Options',
                   'BigBasket_Title', 'BigBasket_Price', 'BigBasket_Original_Price', 'BigBasket_Discount',
                   'BigBasket_Pack_Size', 'BigBasket_Dropdown_Prices', 'Hyperpure_Product_Title', 'Hyperpure_Price',
                   'Hyperpure_Category', 'Hyperpure_SUPERSAVER_Information', 'Agmarknet_Commodity', 'Agmarknet_Variety',
                   'Agmarknet_MAX', 'Agmarknet_MIN', 'Agmarknet_Modal']

        if 'Agmarknet' in selected_websites and not st.session_state.stop_scraping:
            st.write("Scraping Agmarknet...")
            agmarknet_data, agmarknet_file = scrape_agmarknet(driver, search_terms)
            agmarknet_data = agmarknet_data.reindex(columns=columns, fill_value='')
            all_data['Agmarknet'] = agmarknet_data
            st.session_state.download_files['Agmarknet'] = agmarknet_file
            st.write(f"Agmarknet data saved to {agmarknet_file}")

        if 'BigBasket' in selected_websites and not st.session_state.stop_scraping:
            st.write("Scraping BigBasket...")
            bigbasket_data, bigbasket_file = scrape_bigbasket(driver, search_terms)
            bigbasket_data = bigbasket_data.reindex(columns=columns, fill_value='')
            all_data['BigBasket'] = bigbasket_data
            st.session_state.download_files['BigBasket'] = bigbasket_file
            st.write(f"BigBasket data saved to {bigbasket_file}")

        if 'DMart' in selected_websites and not st.session_state.stop_scraping:
            st.write("Scraping DMart...")
            dmart_data, dmart_file = scrape_dmart(driver, search_terms)
            dmart_data = dmart_data.reindex(columns=columns, fill_value='')
            all_data['DMart'] = dmart_data
            st.session_state.download_files['DMart'] = dmart_file
            st.write(f"DMart data saved to {dmart_file}")

        if 'Hyperpure' in selected_websites and not st.session_state.stop_scraping:
            st.write("Scraping Hyperpure...")
            hyperpure_data, hyperpure_file = scrape_hyperpure(driver, search_terms)
            hyperpure_data = hyperpure_data.reindex(columns=columns, fill_value='')
            all_data['Hyperpure'] = hyperpure_data
            st.session_state.download_files['Hyperpure'] = hyperpure_file
            st.write(f"Hyperpure data saved to {hyperpure_file}")

        if 'JioMart' in selected_websites and not st.session_state.stop_scraping:
            st.write("Scraping JioMart...")
            jiomart_data, jiomart_file = scrape_jiomart(driver, search_terms)
            jiomart_data = jiomart_data.reindex(columns=columns, fill_value='')
            all_data['JioMart'] = jiomart_data
            st.session_state.download_files['JioMart'] = jiomart_file
            st.write(f"JioMart data saved to {jiomart_file}")

        # Combine all data into a master DataFrame
        if all_data and not st.session_state.stop_scraping:
            master_data = pd.concat(all_data.values(), ignore_index=True)
            master_output_file = os.path.join(output_folder, 'master_output_for_all.xlsx')
            append_to_excel(master_data, master_output_file)
            st.session_state.download_files['Master'] = master_output_file
            st.success("Data scraping completed successfully!")

    finally:
        # Close the WebDriver
        st.write("Closing WebDriver...")
        driver.quit()

    # Display download buttons for all available files
    if st.session_state.download_files:
        st.write("Download available files:")
        for website, file_path in st.session_state.download_files.items():
            with open(file_path, 'rb') as f:
                st.download_button(
                    label=f"Download {website} Data",
                    data=f,
                    file_name=f'{website}_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key=f'{website}_download_button_post_scrape_{website}'
                )


# Streamlit UI
st.title("Web Scraper")

# Upload the Master_List.xlsx file
uploaded_file = st.file_uploader("Upload your Master_List.xlsx file", type="xlsx")

if uploaded_file is not None:
    # Read the uploaded file into a DataFrame
    df = pd.read_excel(uploaded_file)
    search_terms = df['Vegetables'].tolist()

    # Dropdown menu for selecting websites
    selected_websites = st.multiselect(
        "Select websites to scrape:",
        ['Agmarknet', 'BigBasket', 'DMart', 'Hyperpure', 'JioMart'],
        default=['Agmarknet', 'BigBasket', 'DMart', 'Hyperpure', 'JioMart']
    )

    # Buttons to start and stop scraping
    start_button = st.button("Start Scraping")
    stop_button = st.button("Stop Scraping")

    if start_button:
        # Clear previous downloads
        clear_previous_data()
        # Run the scraping process
        main(selected_websites, search_terms)

    if stop_button:
        st.session_state.stop_scraping = True
        st.warning("Stopping the scraping process...")

    # Display download buttons for all available files
    if st.session_state.download_files:
        st.write("Download available files:")
        for website, file_path in st.session_state.download_files.items():
            with open(file_path, 'rb') as f:
                # Ensure unique keys by appending the website name to the key
                st.download_button(
                    label=f"Download {website} Data",
                    data=f,
                    file_name=f'{website}_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key=f'{website}_download_button_{website}_{int(time.time())}'
                )

else:
    st.warning("Please upload the Master_List.xlsx file to proceed.")
